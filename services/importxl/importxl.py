import openpyxl
import datetime

from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

"""Настройки ексель файла"""
from config.exelconfig import title, width_column
from adminpanel.models import Orders, Product


class Importxl:
    """Импорт данных в эксель"""

    def importxl(orders) -> None:
        """Метод экспорта данных в эксель"""

        book = openpyxl.Workbook()
        sheet = book.active

        fontStyle = Font(size="10")
        fontStyletitle = Font(size="10", bold=True)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        """задаем ширину ячеек"""
        for object in width_column:
            sheet.column_dimensions[object['column']].width = object['width']

        """Задаем стили титульной строки"""
        for counter, item in enumerate(title, 1):
            sheet.cell(row=1, column=counter).value = item
            sheet.cell(row=1, column=counter).font = fontStyletitle
            sheet.cell(row=1, column=counter).fill = PatternFill(
                start_color="ffff00",
                end_color="ffff00",
                fill_type="solid")

            currentCell = sheet.cell(row=1, column=counter)
            currentCell.alignment = Alignment(horizontal='center')
            counter += 1

        """записываем данные в таблицу"""
        counter = 2
        for obj in orders:
            products = Product.objects.filter(number_order=obj.number)

            sheet.cell(row=counter, column=1).value = obj.number
            sheet.cell(row=counter, column=2).value = obj.id_manager
            sheet.cell(row=counter, column=3).value = obj.data_orders.strftime(
                "%d-%m-%y")

            for i, product in enumerate(products, 0):
                if i == 0:
                    Importxl.color_column(sheet, counter, 4, product.status)
                    sheet.cell(row=counter, column=4).value = product.status
                    sheet.cell(
                        row=counter, column=5).value = product.descriptions
                    sheet.cell(row=counter, column=6).value = product.brand
                    sheet.cell(row=counter, column=7).value = product.article
                    sheet.cell(row=counter, column=8).value = product.quantity
                    sheet.cell(
                        row=counter, column=9).value = product.price_product

                    sheet.cell(row=counter, column=10).value = obj.price
                    sheet.cell(row=counter, column=11).value = obj.paid
                    sheet.cell(row=counter, column=12).value = obj.debt
                    sheet.cell(row=counter, column=13).value = str(
                        obj.name_client)
                    sheet.cell(row=counter, column=14).value = obj.phone

                    sheet.cell(row=counter, column=15).value = product.date_deadline.strftime(
                        "%d-%m-%y")
                    sheet.cell(
                        row=counter, column=16).value = product.distributor
                    sheet.cell(
                        row=counter, column=17).value = product.order_distributer
                    sheet.cell(row=counter, column=18).value = product.comment

                    sheet.cell(row=counter, column=19).value = obj.update_date.strftime("%d-%m-%y")
                    sheet.cell(row=counter, column=20).value = obj.manager

                else:
                    Importxl.color_column(sheet, counter, 4, product.status)
                    sheet.cell(row=counter, column=4).value = product.status
                    sheet.cell(
                        row=counter, column=5).value = product.descriptions
                    sheet.cell(row=counter, column=6).value = product.brand
                    sheet.cell(row=counter, column=7).value = product.article
                    sheet.cell(row=counter, column=8).value = product.quantity
                    sheet.cell(
                        row=counter, column=9).value = product.price_product

                    sheet.cell(row=counter, column=15).value = product.date_deadline.strftime("%d-%m-%y")
                    sheet.cell(
                        row=counter, column=16).value = product.distributor
                    sheet.cell(
                        row=counter, column=17).value = product.order_distributer
                    sheet.cell(row=counter, column=18).value = product.comment

                for item in range(1, 21):
                    sheet.cell(row=counter, column=item).border = thin_border
                    sheet.cell(row=counter, column=item).font = fontStyle
                counter += 1

        sheet.row_dimensions[20]

        #пути сохранения сервр-локалка не трогать
        #book.save('/home/sammy/myprojectdir/media/myexel.xlsx')
        book.save('media/myexel.xlsx')
        book.close()

    def color_column(sheet, rows, columns, status) -> None:
        """Красим ячейку в зависимости от статуса"""

        if status == 'Отгружен поставщиком':
            sheet.cell(row=rows, column=4).fill = PatternFill(
                start_color="fdf731",
                fill_type="solid")
        if status == 'Заказан поставщику':
            sheet.cell(row=rows, column=columns).fill = PatternFill(
                start_color="fdf731",
                fill_type="solid")
        if status == 'Требует внимания':
            sheet.cell(row=rows, column=columns).fill = PatternFill(
                start_color="FF9C59",
                fill_type="solid")
        if status == 'Отказ поставщика':
            sheet.cell(row=rows, column=columns).fill = PatternFill(
                start_color="E93F33",
                fill_type="solid")
        if status == 'Перезаказ поставщику':
            sheet.cell(row=rows, column=columns).fill = PatternFill(
                start_color="C7C227",
                fill_type="solid")
        if status == 'Пришел на склад':
            sheet.cell(row=rows, column=columns).fill = PatternFill(
                start_color="76f9fb",
                fill_type="solid")
        if status == 'Готов к выдаче':
            sheet.cell(row=rows, column=columns).fill = PatternFill(
                start_color="e55ffe",
                fill_type="solid")
        if status == 'Выдан':
            sheet.cell(row=rows, column=columns).fill = PatternFill(
                start_color="7ff21a",
                fill_type="solid")
        if status == 'Отказ клиента':
            sheet.cell(row=rows, column=columns).fill = PatternFill(
                start_color="85241E",
                fill_type="solid")
